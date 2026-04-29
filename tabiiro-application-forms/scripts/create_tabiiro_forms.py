import csv
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import xml.etree.ElementTree as ET
from datetime import date, timedelta
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

DEFAULT_PLAN_CSV = Path(r"C:\Users\NX023066\Downloads\【BM】電子雑誌・AJ営業部／【公式】プラン一覧.xlsx")
DEFAULT_MAIN_TEMPLATE = Path(r"C:\Users\NX023066\Downloads\【A3横】旅色／申込書（ver.004）_260303.xlsx")
DEFAULT_OPTION_TEMPLATE = Path(r"C:\Users\NX023066\Downloads\【A3横】旅色／オプションサービス申込書（ver.013）_260303.xlsx")
DEFAULT_OUTPUT_ROOT = Path(r"C:\Users\NX023066\Documents\New project\outputs")

ASSIGNEE = "渡邊裕介"
DISCOUNT_LABEL = "貴社向け特別値引き"
SHINCHAKU_TAX_IN = 220_000

JP_HOLIDAYS_2026 = {
    date(2026, 1, 1), date(2026, 1, 12), date(2026, 2, 11), date(2026, 2, 23),
    date(2026, 3, 20), date(2026, 4, 29), date(2026, 5, 3), date(2026, 5, 4),
    date(2026, 5, 5), date(2026, 5, 6), date(2026, 7, 20), date(2026, 8, 11),
    date(2026, 9, 21), date(2026, 9, 22), date(2026, 9, 23), date(2026, 10, 12),
    date(2026, 11, 3), date(2026, 11, 23),
}

UNSPECIFIED_ZERO_CELLS_MAIN = (
    "BF23", "BS23", "BF33", "BS33", "BF43", "BS43",
    "BF53", "BS53", "BF63", "BS63", "BF73", "BS73",
)
UNSPECIFIED_ZERO_CELLS_OPTION = (
    "BF16", "BF56", "BS56", "BF70", "BS70", "S46", "AF46", "S73", "AF73",
)

PR_ALIASES = {
    "旅色PR広告【記事上部】": "記事枠3w",
    "記事上部": "記事枠3w",
}

HP_ALIASES = {
    "あり": "簡易HP_S_ssl",
    "有": "簡易HP_S_ssl",
    "true": "簡易HP_S_ssl",
    "1": "簡易HP_S_ssl",
    "HPS_ssl": "簡易HP_S_ssl",
    "簡易HPS_ssl": "簡易HP_S_ssl",
    "簡易HP_S_ssl": "簡易HP_S_ssl",
}


def parse_amount(value) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).translate(str.maketrans("０１２３４５６７８９．，", "0123456789.,"))
    text = text.replace(",", "").replace("円", "").replace("税込", "").replace("税抜", "").strip()
    if text in {"", "無料", "-", "なし"}:
        return 0
    multiplier = 1
    if "万" in text:
        multiplier = 10_000
        text = text.replace("万円", "").replace("万", "").strip()
    return int(round(float(text) * multiplier))


def truthy(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"あり", "有", "true", "yes", "y", "1", "する"}


def specified(value) -> bool:
    if value is None:
        return False
    text = str(value).strip().lower()
    return text not in {"", "なし", "無", "false", "no", "n", "0", "null"}


def parse_month(value, default_year=2026):
    if isinstance(value, dict):
        return int(value.get("year", default_year)), int(value["month"])
    nums = [int(n) for n in re.findall(r"\d+", str(value))]
    if len(nums) >= 2 and nums[0] > 1900:
        return nums[0], nums[1]
    if nums:
        return default_year, nums[0]
    raise ValueError(f"掲載月を解釈できません: {value}")


def add_months(year: int, month: int, months: int = 1) -> tuple[int, int]:
    month_index = (year * 12 + (month - 1)) + months
    return month_index // 12, month_index % 12 + 1


def is_business_day(d: date) -> bool:
    return d.weekday() < 5 and d not in JP_HOLIDAYS_2026


def next_business_day(d: date) -> date:
    while not is_business_day(d):
        d += timedelta(days=1)
    return d


def previous_business_day(d: date) -> date:
    while not is_business_day(d):
        d -= timedelta(days=1)
    return d


def last_business_day(year: int, month: int) -> date:
    d = date(year, 12, 31) if month == 12 else date(year, month + 1, 1) - timedelta(days=1)
    return previous_business_day(d)


def main_start_date(year: int, month: int) -> date:
    return next_business_day(date(year, month, 25))


def taiwan_start_date(year: int, month: int) -> date:
    return previous_business_day(last_business_day(year, month) - timedelta(days=1))


def hp_start_date(year: int, month: int) -> date:
    return next_business_day(date(year + (1 if month == 12 else 0), 1 if month == 12 else month + 1, 25))


def parse_date_text(value, fallback_year=2026) -> date:
    nums = [int(n) for n in re.findall(r"\d+", str(value))]
    if len(nums) >= 3:
        return date(nums[0], nums[1], nums[2])
    if len(nums) >= 2:
        return date(fallback_year, nums[0], nums[1])
    raise ValueError(f"日付を解釈できません: {value}")


def load_plan_rows(path: Path):
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            # Some source workbooks expose only A1 through openpyxl despite opening normally in Excel.
            # In that case, fall back to Excel COM via PowerShell and re-read a temporary CSV.
            if all(wb[s].max_row <= 1 and wb[s].max_column <= 1 for s in ("旅色_国内版", "旅色_多言語独自ページ制作") if s in wb.sheetnames):
                tmp_csv = export_plan_workbook_to_csv(path)
                return load_plan_rows(tmp_csv)
            rows = []
            priority = ["旅色_国内版", "旅色_多言語独自ページ制作", "旅色（台湾版）"]
            sheet_names = [s for s in priority if s in wb.sheetnames]
            sheet_names += [s for s in wb.sheetnames if s not in sheet_names]
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    rows.append(["" if v is None else str(v) for v in row])
            return rows
        finally:
            wb.close()
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.reader(f))


def export_plan_workbook_to_csv(path: Path) -> Path:
    out = Path(tempfile.gettempdir()) / "tabiiro_plan_list_combined.csv"
    ps = rf'''
function CsvEscape([string]$s) {{
  if($null -eq $s) {{ return '' }}
  $s=$s -replace '"','""'
  if($s -match '[,`r`n"]') {{ return '"' + $s + '"' }}
  return $s
}}
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$writer = $null
try {{
  $wb = $excel.Workbooks.Open('{str(path)}', 0, $true)
  $writer = [System.IO.StreamWriter]::new('{str(out)}', $false, [System.Text.UTF8Encoding]::new($true))
  foreach($sheetName in @('旅色_国内版','旅色_多言語独自ページ制作')) {{
    $ws = $wb.Worksheets.Item($sheetName)
    $range = $ws.UsedRange
    $vals = $range.Value2
    $rows = $range.Rows.Count
    $cols = $range.Columns.Count
    for($r=1; $r -le $rows; $r++) {{
      $fields = New-Object System.Collections.Generic.List[string]
      for($c=1; $c -le $cols; $c++) {{
        $fields.Add((CsvEscape ([string]$vals[$r,$c])))
      }}
      $writer.WriteLine(($fields -join ','))
    }}
  }}
  $writer.Close()
  $wb.Close($false)
}} finally {{
  if($writer) {{ $writer.Dispose() }}
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}}
'''
    subprocess.run(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps], check=True)
    return out


def normalize_plan(value: str) -> str:
    return re.sub(r"\s+", "", str(value)).upper()


def find_plan(rows, plan_name: str):
    exact = str(plan_name).strip()
    target = normalize_plan(plan_name)
    category = ""
    fallback = None
    for row in rows[9:]:
        if len(row) > 1 and row[1].strip():
            category = row[1].strip()
        if len(row) > 2 and row[2].strip() == exact:
            return row, category
        if len(row) > 2 and normalize_plan(row[2]) == target and fallback is None:
            fallback = (row, category)
    if fallback:
        return fallback
    raise ValueError(f"プラン一覧に見つかりません: {plan_name}")


def plan_tax_in_without_shinchaku(row) -> int:
    for idx in (11, 13):
        if len(row) > idx and str(row[idx]).strip():
            return parse_amount(row[idx])
    return 0


def plan_total_tax_in(row) -> int:
    return parse_amount(row[13] if len(row) > 13 else 0)


def plan_payment_count(row) -> int:
    return parse_amount(row[21] if len(row) > 21 else 1) or 1


def plan_months(row, plan_name: str) -> int:
    if len(row) > 6 and str(row[6]).strip():
        m = re.search(r"\d+", str(row[6]))
        if m:
            return int(m.group())
    name = normalize_plan(plan_name)
    if name.endswith("A"):
        return 12
    if name.endswith("B"):
        return 24
    if name.endswith("C"):
        return 36
    return 1


def months_from_plan_suffix(plan_name: str) -> int:
    name = normalize_plan(plan_name)
    if name.endswith("A"):
        return 12
    if name.endswith("B"):
        return 24
    if name.endswith("C"):
        return 36
    return 1


def service_amount_count(rows, plan_name: str):
    row, _ = find_plan(rows, plan_name)
    return plan_tax_in_without_shinchaku(row) or plan_total_tax_in(row), plan_payment_count(row)


def normalize_hp_plan(value) -> str:
    text = str(value).replace("。", " ").replace("、", " ").strip()
    text = re.sub(r"\s*\d+月.*$", "", text).strip()
    return HP_ALIASES.get(text, text)


def unprotect_workbook(wb):
    if wb.security is not None:
        wb.security.lockStructure = False
        wb.security.lockWindows = False
    for ws in wb.worksheets:
        ws.protection.sheet = False
        ws.protection.objects = False
        ws.protection.scenarios = False


def strip_workbook_xml(path: Path, *, remove_protection: bool = True, remove_phonetics: bool = True):
    tmp = path.with_suffix(".tmp.xlsx")
    with ZipFile(path, "r") as zin, ZipFile(tmp, "w", ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith(".xml"):
                text = data.decode("utf-8", errors="ignore")
                if remove_protection:
                    if item.filename.startswith("xl/worksheets/"):
                        text = re.sub(r"<sheetProtection\b[^>]*/>", "", text)
                    elif item.filename == "xl/workbook.xml":
                        text = re.sub(r"<workbookProtection\b[^>]*/>", "", text)
                if remove_phonetics:
                    text = re.sub(r"<rPh\b[^>]*>.*?</rPh>", "", text, flags=re.DOTALL)
                    text = re.sub(r"<phoneticPr\b[^>]*/>", "", text)
                    text = re.sub(r"<phoneticPr\b[^>]*>.*?</phoneticPr>", "", text, flags=re.DOTALL)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    os.replace(tmp, path)


def sheet_xml_path(zipped_workbook: ZipFile, sheet_name: str) -> str:
    ns_main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    ns_rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    workbook = ET.fromstring(zipped_workbook.read("xl/workbook.xml"))
    rels = ET.fromstring(zipped_workbook.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    for sheet in workbook.find(ns_main + "sheets"):
        if sheet.attrib.get("name") == sheet_name:
            target = rid_to_target[sheet.attrib[ns_rel + "id"]]
            return "xl/" + target.lstrip("/")
    raise KeyError(f"Sheet not found: {sheet_name}")


def extract_cols_xml(sheet_xml: str) -> str:
    match = re.search(r"<cols>.*?</cols>", sheet_xml, flags=re.DOTALL)
    return match.group(0) if match else ""


def replace_cols_xml(sheet_xml: str, cols_xml: str) -> str:
    sheet_xml = re.sub(r"<cols>.*?</cols>", "", sheet_xml, flags=re.DOTALL)
    if not cols_xml:
        return sheet_xml
    return sheet_xml.replace("<sheetData>", cols_xml + "<sheetData>", 1)


def restore_column_widths_xml(template_path: Path, generated_path: Path, sheet_name: str):
    with ZipFile(template_path, "r") as template_zip:
        template_sheet_path = sheet_xml_path(template_zip, sheet_name)
        template_cols = extract_cols_xml(template_zip.read(template_sheet_path).decode("utf-8"))

    tmp = generated_path.with_suffix(".layout.xlsx")
    with ZipFile(generated_path, "r") as zin, ZipFile(tmp, "w", ZIP_DEFLATED) as zout:
        output_sheet_path = sheet_xml_path(zin, sheet_name)
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == output_sheet_path:
                text = data.decode("utf-8")
                data = replace_cols_xml(text, template_cols).encode("utf-8")
            zout.writestr(item, data)
    os.replace(tmp, generated_path)


def merged_non_top_left(ws):
    skip = set()
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                skip.add((row, col))
    return skip


def jsonable_cell_value(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def collect_value_changes(template_path: Path, generated_path: Path, sheet_name: str):
    template_wb = openpyxl.load_workbook(template_path, data_only=False)
    generated_wb = openpyxl.load_workbook(generated_path, data_only=False)
    try:
        template_ws = template_wb[sheet_name]
        generated_ws = generated_wb[sheet_name]
        skip = merged_non_top_left(template_ws)
        max_row = max(template_ws.max_row, generated_ws.max_row)
        max_col = max(template_ws.max_column, generated_ws.max_column)
        changes = []
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                if (row, col) in skip:
                    continue
                template_value = template_ws.cell(row, col).value
                generated_value = generated_ws.cell(row, col).value
                if template_value != generated_value:
                    changes.append(
                        {
                            "cell": f"{openpyxl.utils.get_column_letter(col)}{row}",
                            "value": jsonable_cell_value(generated_value),
                        }
                    )
        return changes
    finally:
        template_wb.close()
        generated_wb.close()


def restore_template_layout(template_path: Path, generated_path: Path, sheet_name: str):
    generated_values = Path(tempfile.gettempdir()) / f"{generated_path.stem}_values.xlsx"
    changes_path = Path(tempfile.gettempdir()) / f"{generated_path.stem}_changes.json"
    shutil.copy2(generated_path, generated_values)
    changes = collect_value_changes(template_path, generated_values, sheet_name)

    shutil.copy2(template_path, generated_path)
    strip_workbook_xml(generated_path, remove_protection=True, remove_phonetics=True)
    changes_path.write_text(
        json.dumps({"path": str(generated_path), "sheet": sheet_name, "changes": changes}, ensure_ascii=True),
        encoding="utf-8",
    )

    ps = rf'''
$ErrorActionPreference = 'Stop'
$data = Get-Content -LiteralPath '{changes_path}' -Encoding UTF8 -Raw | ConvertFrom-Json
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
  $wb = $excel.Workbooks.Open($data.path)
  try {{
    $ws = $wb.Worksheets.Item($data.sheet)
    try {{ $ws.Unprotect() | Out-Null }} catch {{}}
    try {{ $wb.Unprotect() | Out-Null }} catch {{}}
    foreach ($change in $data.changes) {{
      $cell = $ws.Range($change.cell)
      $target = if ($cell.MergeCells) {{ $cell.MergeArea.Cells.Item(1,1) }} else {{ $cell }}
      if ($null -eq $change.value -or [string]$change.value -eq '') {{
        if ($target.MergeCells) {{ $target.MergeArea.ClearContents() | Out-Null }} else {{ $target.ClearContents() | Out-Null }}
      }} else {{
        try {{
          $target.Value2 = $change.value
        }} catch {{
          $target.Value2 = [string]$change.value
        }}
      }}
    }}
    try {{ $ws.UsedRange.Phonetics.Delete() | Out-Null }} catch {{}}
    try {{ $ws.UsedRange.Phonetics.Visible = $false }} catch {{}}
    $wb.Save()
  }} finally {{
    $wb.Close($true)
  }}
}} finally {{
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}}
'''
    try:
        subprocess.run(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps], check=True)
        strip_workbook_xml(generated_path, remove_protection=True, remove_phonetics=True)
        restore_column_widths_xml(template_path, generated_path, sheet_name)
    finally:
        generated_values.unlink(missing_ok=True)
        changes_path.unlink(missing_ok=True)


def allocate_amount(acquired: int, main_list: int, has_shinchaku: bool):
    if acquired <= 0:
        return 0, 0
    if not has_shinchaku:
        return min(acquired, main_list), 0
    shinchaku_paid = min(acquired, SHINCHAKU_TAX_IN)
    main_paid = min(max(acquired - shinchaku_paid, 0), main_list)
    return main_paid, shinchaku_paid


def set_date(ws, cells, d: date):
    y_cell, m_cell, d_cell = cells
    ws[y_cell] = d.year - 2000
    ws[m_cell] = d.month
    ws[d_cell] = d.day


def set_discount_block(ws, label_cell, tax_cell, mark_cell, discount_cell, yen_cell, list_amount, paid_amount):
    discount = max(list_amount - paid_amount, 0)
    if discount:
        ws[label_cell] = DISCOUNT_LABEL
        ws[tax_cell] = "（税込）"
        ws[mark_cell] = "▲"
        ws[discount_cell] = discount
        ws[yen_cell] = "円"
    else:
        for cell in (label_cell, tax_cell, mark_cell, discount_cell, yen_cell):
            ws[cell] = None


def per_payment(paid_amount: int, count: int) -> int:
    return int(round(paid_amount / count)) if count else paid_amount


def clear_cells(ws, cells):
    for cell in cells:
        ws[cell] = None


def fill_service_block(ws, *, date_cells, months_cell, facilities_cell, list_cell,
                       discount_cells, total_cell, count_cell, per_cell,
                       start: date, months: int, facilities: int,
                       list_amount: int, paid_amount: int, payment_count: int):
    set_date(ws, date_cells, start)
    ws[months_cell] = months
    ws[facilities_cell] = facilities
    ws[list_cell] = list_amount
    set_discount_block(ws, *discount_cells, list_amount, paid_amount)
    ws[total_cell] = paid_amount
    if paid_amount > 0:
        ws[count_cell] = payment_count
        ws[per_cell] = per_payment(paid_amount, payment_count)
    else:
        ws[count_cell] = None
        ws[per_cell] = 0


def parse_payment_start(value, default_year=2026):
    if not value:
        return None
    nums = [int(n) for n in re.findall(r"\d+", str(value))]
    if len(nums) >= 2 and nums[0] > 1900:
        return nums[0] - 2000, nums[1]
    if nums:
        return default_year - 2000, nums[0]
    return None


def set_payment_start(ws, year_cell: str, month_cell: str, payment_start):
    if not payment_start:
        return
    yy, mm = payment_start
    ws[year_cell] = yy
    ws[month_cell] = mm


def fill_main_form(data: dict, rows, out_dir: Path, year: int, month: int) -> Path:
    wb = openpyxl.load_workbook(Path(data.get("main_template", DEFAULT_MAIN_TEMPLATE)))
    ws = wb["ver.004"]
    clear_cells(ws, UNSPECIFIED_ZERO_CELLS_MAIN)

    plan_name = data["本誌プラン名"]
    plan_row, _ = find_plan(rows, plan_name)
    has_shinchaku = truthy(data.get("新着", data.get("新着プラン名")))
    taiwan = truthy(data.get("台湾案件", data.get("台湾")))
    facilities = parse_amount(data.get("施設数", 1)) or 1

    acquired = parse_amount(data.get("獲得金額税込", data.get("獲得金額（税込）", data.get("獲得金額", 0))))
    if "獲得金額税抜" in data:
        acquired = int(round(parse_amount(data["獲得金額税抜"]) * 1.1))

    main_list = plan_tax_in_without_shinchaku(plan_row)
    main_paid, shinchaku_paid = allocate_amount(acquired, main_list, has_shinchaku)
    payment = str(data.get("支払方法", ""))
    payment_start = parse_payment_start(data.get("支払い開始月", data.get("支払開始月")), year)
    main_count = plan_payment_count(plan_row)

    fill_service_block(
        ws,
        date_cells=("AT5", "AW5", "AZ5"),
        months_cell="BD5",
        facilities_cell="BH5",
        list_cell="BF7",
        discount_cells=("AO11", "BB11", "BE11", "BF11", "BK11"),
        total_cell="BF13",
        count_cell="BQ12",
        per_cell="BS13",
        start=main_start_date(year, month),
        months=plan_months(plan_row, plan_name),
        facilities=facilities,
        list_amount=main_list,
        paid_amount=main_paid,
        payment_count=main_count,
    )
    if main_paid > 0 and "口座" in payment:
        ws["BP8"] = "☑"
        set_payment_start(ws, "BQ10", "BS10", payment_start)

    if has_shinchaku:
        fill_service_block(
            ws,
            date_cells=("AT17", "AW17", "AZ17"),
            months_cell="BD17",
            facilities_cell="BH17",
            list_cell="BF19",
            discount_cells=("AO21", "BB21", "BE21", "BF21", "BK21"),
            total_cell="BF23",
            count_cell="BQ22",
            per_cell="BS23",
            start=last_business_day(year, month),
            months=1,
            facilities=facilities,
            list_amount=SHINCHAKU_TAX_IN,
            paid_amount=shinchaku_paid,
            payment_count=main_count,
        )
        if shinchaku_paid > 0 and "口座" in payment:
            ws["BP19"] = "☑"
            set_payment_start(ws, "BQ20", "BS20", payment_start)

    if truthy(data.get("入稿代行")):
        nyuko_plan = data.get("入稿代行プラン名", "TG_入稿a~C")
        nyuko_amount, nyuko_count = service_amount_count(rows, nyuko_plan)
        fill_service_block(
            ws,
            date_cells=("AT27", "AW27", "AZ27"),
            months_cell="BD27",
            facilities_cell="BH27",
            list_cell="BF29",
            discount_cells=("AO31", "BB31", "BE31", "BF31", "BK31"),
            total_cell="BF33",
            count_cell="BQ32",
            per_cell="BS33",
            start=date(year, month, 1),
            months=1,
            facilities=facilities,
            list_amount=nyuko_amount,
            paid_amount=parse_amount(data.get("入稿代行支払税込", 0)),
            payment_count=nyuko_count,
        )
        if parse_amount(data.get("入稿代行支払税込", 0)) > 0 and "口座" in payment:
            ws["BP29"] = "☑"
            set_payment_start(ws, "BQ30", "BS30", payment_start)

    if taiwan:
        taiwan_plan = data.get("台湾プラン名", data.get("多言語プラン名", "twA"))
        taiwan_list, taiwan_count_from_plan = service_amount_count(rows, taiwan_plan)
        override_taiwan_list = parse_amount(data.get("台湾定価税込", data.get("多言語定価税込", 0)))
        if override_taiwan_list:
            taiwan_list = override_taiwan_list
        taiwan_paid = parse_amount(data.get("台湾支払税込", data.get("多言語支払税込", 0)))
        taiwan_count = parse_amount(data.get("台湾支払回数", data.get("多言語支払回数", taiwan_count_from_plan))) or 1
        ws["AO59"] = "☑"  # 繁体字
        ws["AS59"] = "□"
        fill_service_block(
            ws,
            date_cells=("AT57", "AW57", "AZ57"),
            months_cell="BD57",
            facilities_cell="BH57",
            list_cell="BF59",
            discount_cells=("AO61", "BB61", "BE61", "BF61", "BK61"),
            total_cell="BF63",
            count_cell="BQ62",
            per_cell="BS63",
            start=taiwan_start_date(year, month),
            months=months_from_plan_suffix(taiwan_plan),
            facilities=facilities,
            list_amount=taiwan_list,
            paid_amount=taiwan_paid,
            payment_count=taiwan_count,
        )
        if parse_amount(data.get("台湾支払税込", data.get("多言語支払税込", 0))) > 0 and "口座" in payment:
            ws["BP59"] = "☑"
            set_payment_start(ws, "BQ60", "BS60", payment_start)
        if has_shinchaku:
            taiwan_new_paid = parse_amount(data.get("台湾新着支払税込", data.get("多言語新着支払税込", 0)))
            fill_service_block(
                ws,
                date_cells=("AT67", "AW67", "AZ67"),
                months_cell="BD67",
                facilities_cell="BH67",
                list_cell="BF69",
                discount_cells=("AO71", "BB71", "BE71", "BF71", "BK71"),
                total_cell="BF73",
                count_cell="BQ72",
                per_cell="BS73",
                start=last_business_day(year, month),
                months=1,
                facilities=facilities,
                list_amount=SHINCHAKU_TAX_IN,
                paid_amount=taiwan_new_paid,
                payment_count=taiwan_count,
            )
            if taiwan_new_paid > 0 and "口座" in payment:
                ws["BP69"] = "☑"
                set_payment_start(ws, "BQ70", "BS70", payment_start)

    if specified(data.get("HP")):
        hp_value = data.get("HPプラン名") or data.get("HP")
        hp_plan = normalize_hp_plan(hp_value)
        hp_amount, hp_count = service_amount_count(rows, hp_plan)
        if data.get("HP掲載月") or data.get("HP開始月"):
            hp_year, hp_month = parse_month(data.get("HP掲載月", data.get("HP開始月")), year)
        else:
            hp_year, hp_month = add_months(year, month)
        start = next_business_day(date(hp_year, hp_month, 25))
        ws["AO39"] = "☑" if "S" in hp_plan.upper() else "□"
        ws["AS39"] = "☑" if "C" in hp_plan.upper() else "□"
        fill_service_block(
            ws,
            date_cells=("AT37", "AW37", "AZ37"),
            months_cell="BD37",
            facilities_cell="BH37",
            list_cell="BF39",
            discount_cells=("AO41", "BB41", "BE41", "BF41", "BK41"),
            total_cell="BF43",
            count_cell="BQ42",
            per_cell="BS43",
            start=start,
            months=12,
            facilities=facilities,
            list_amount=hp_amount,
            paid_amount=parse_amount(data.get("HP支払税込", 0)),
            payment_count=hp_count,
        )
        if parse_amount(data.get("HP支払税込", 0)) > 0 and "口座" in payment:
            ws["BP39"] = "☑"
            set_payment_start(ws, "BQ40", "BS40", payment_start)

    ws["AW76"] = ASSIGNEE
    ws["BQ76"] = plan_name

    if data.get("申込日"):
        d = parse_date_text(data["申込日"], year)
        set_date(ws, ("H6", "K6", "N6"), d)
    if data.get("備考入力"):
        ws["AS78"] = str(data["備考入力"])
    if data.get("特記事項入力"):
        ws["AS81"] = str(data["特記事項入力"])

    safe_plan = re.sub(r"[^\w一-龥ぁ-んァ-ンー]+", "_", plan_name)
    out = out_dir / f"旅色_申込書_{safe_plan}_{year}年{month}月.xlsx"
    unprotect_workbook(wb)
    wb.save(out)
    restore_template_layout(Path(data.get("main_template", DEFAULT_MAIN_TEMPLATE)), out, "ver.004")
    return out


def create_option_form(data: dict, rows, out_dir: Path, year: int, month: int) -> Path:
    wb = openpyxl.load_workbook(Path(data.get("option_template", DEFAULT_OPTION_TEMPLATE)))
    ws = wb["ver.013"]
    clear_cells(ws, UNSPECIFIED_ZERO_CELLS_OPTION)

    pr = data.get("PR記事")
    if isinstance(pr, dict):
        service_label = pr.get("種類") or pr.get("サービス名") or pr.get("プラン名")
        start_text = pr.get("開始日") or data.get("PR開始日")
        paid_amount = parse_amount(pr.get("支払税込", data.get("PR支払税込", 0)))
    else:
        service_label = str(pr)
        start_text = data.get("PR開始日")
        paid_amount = parse_amount(data.get("PR支払税込", 0))

    if not service_label or service_label in {"あり", "有", "true"}:
        raise ValueError("PR記事ありの場合は、PR記事の種類/サービス名を指定してください。")
    if not start_text:
        raise ValueError("PR記事ありの場合は、開始日を指定してください。")

    lookup_plan = PR_ALIASES.get(service_label, service_label)
    list_amount, _ = service_amount_count(rows, lookup_plan)
    start = parse_date_text(start_text, year)

    ws["AS62"] = service_label
    ws["AU64"] = start.year - 2000
    ws["AX64"] = start.month
    ws["BD64"] = 3
    ws["BH64"] = parse_amount(data.get("施設数", 1)) or 1
    ws["BF66"] = list_amount
    set_discount_block(ws, "AO68", "AY68", "BE68", "BF68", "BJ68", list_amount, paid_amount)
    ws["BF70"] = paid_amount
    if paid_amount > 0:
        ws["BR69"] = 1
    ws["BS70"] = paid_amount
    ws["AW74"] = ASSIGNEE

    safe = re.sub(r"[^\w一-龥ぁ-んァ-ンー]+", "_", service_label)
    out = out_dir / f"旅色_オプションサービス申込書_{safe}_{year}年{month}月.xlsx"
    unprotect_workbook(wb)
    wb.save(out)
    restore_template_layout(Path(data.get("option_template", DEFAULT_OPTION_TEMPLATE)), out, "ver.013")
    return out


def fill_application(data: dict):
    year, month = parse_month(data["掲載月"], data.get("年", 2026))
    rows = load_plan_rows(Path(data.get("plan_csv", DEFAULT_PLAN_CSV)))
    safe_plan = re.sub(r"[^\w一-龥ぁ-んァ-ンー]+", "_", data["本誌プラン名"])
    out_dir = Path(data.get("output_root", DEFAULT_OUTPUT_ROOT)) / f"tabiiro_{year}{month:02d}_{safe_plan}"
    out_dir.mkdir(parents=True, exist_ok=True)

    outputs = [fill_main_form(data, rows, out_dir, year, month)]
    if specified(data.get("PR記事")):
        outputs.append(create_option_form(data, rows, out_dir, year, month))
    return outputs


def read_input():
    raw_bytes = sys.stdin.buffer.read()
    raw = ""
    if raw_bytes:
        for encoding in ("utf-8-sig", "cp932", "utf-16"):
            try:
                raw = raw_bytes.decode(encoding).strip()
                break
            except UnicodeDecodeError:
                pass
    if not raw and len(sys.argv) > 1:
        raw = Path(sys.argv[1]).read_text(encoding="utf-8-sig")
    if not raw:
        raise SystemExit("JSON input is required via stdin or file path argument.")
    return json.loads(raw)


def main():
    outputs = fill_application(read_input())
    print(json.dumps({"outputs": [str(p) for p in outputs]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
