import os
from openpyxl import load_workbook
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

downloads = r"C:\Users\NX023066\Downloads"
gdrive = r"G:\マイドライブ\codex-skills\tbiiro-renewal"

search_dirs = [downloads, gdrive]
terms = ["epice", "エピス", "315399"]

print("Starting deep search in Excel files...")

for sdir in search_dirs:
    if not os.path.exists(sdir):
        continue
    for root, dirs, files in os.walk(sdir):
        for f in files:
            if f.endswith(".xlsx") and not f.startswith("~$"):
                path = os.path.join(root, f)
                try:
                    wb = load_workbook(path, read_only=True, data_only=True)
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        # Search for Epice first in this file
                        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                            row_str = " ".join(str(cell) for cell in row if cell is not None)
                            if any(t in row_str.lower() for t in terms):
                                # Found Epice! Let's print the row context
                                print(f"\n[FOUND EPICE] File: {f} | Sheet: {sheet} | Row: {r_idx}")
                                print("  Content:", row_str[:200])
                except Exception as e:
                    pass
