import openpyxl
import sys, io
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

f = Path.home() / 'Downloads' / '【旅色SNS】営業支援投稿（～2024年12月） (1).xlsx'
if not f.exists():
    print(f"File not found: {f}")
    sys.exit(1)

print("Loading workbook (without read_only)...")
wb = openpyxl.load_workbook(f, data_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'\nSearching sheet: "{name}" ...')
    
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_str = ' '.join([str(c) for c in row if c is not None])
        if '鮨' in row_str or 'ふみ' in row_str or '313380' in row_str:
            print(f"  [MATCH] Row {r_idx}: {[str(c) for c in row if c is not None]}")
