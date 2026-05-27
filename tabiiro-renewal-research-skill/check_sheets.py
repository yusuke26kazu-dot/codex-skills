from openpyxl import load_workbook
import sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = load_workbook(r'G:\マイドライブ\codex-skills\tbiiro-renewal\各種素材類・データ\SNS配信\台湾旅色_IG・FB掲載一覧.xlsx', read_only=True)
print("Sheets in Taiwan Excel:")
for sheet in wb.sheetnames:
    print(" -", sheet)
