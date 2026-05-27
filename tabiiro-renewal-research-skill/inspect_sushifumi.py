import openpyxl
import sys, io
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

f = Path.home() / 'Downloads' / '入稿シート_TG3_【鮨ふみ】_2021.4～.xlsx'
if not f.exists():
    print(f"File not found: {f}")
    sys.exit(1)

print("Loading workbook...")
wb = openpyxl.load_workbook(f, data_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'\nSheet: "{name}" ...')
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_str = ' '.join([str(c) for c in row if c is not None])
        if '2026' in row_str or '3/16' in row_str or '3月16日' in row_str or 'fb' in row_str.lower() or 'facebook' in row_str.lower() or 'sns' in row_str.lower():
            print(f"  Row {r_idx}: {[str(c) for c in row if c is not None]}")
