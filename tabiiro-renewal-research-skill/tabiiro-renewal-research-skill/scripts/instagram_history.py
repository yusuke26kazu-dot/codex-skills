#!/usr/bin/env python3
"""Search Tabiiro SNS sales-support Instagram workbooks for posting history."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter, deque
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_WORKBOOK_NAMES = [
    "\u2605\u3010\u65c5\u8272SNS\u3011\u55b6\u696d\u652f\u63f4\u6295\u7a3f.xlsx",
    "\u3010\u65c5\u8272SNS\u3011\u55b6\u696d\u652f\u63f4\u6295\u7a3f.xlsx",
    "\u3010\u65c5\u8272SNS\u3011\u55b6\u696d\u652f\u63f4\u6295\u7a3f\uff08\uff5e2024\u5e7412\u6708\uff09.xlsx",
]

SHEET_INCLUDE_HINTS = (
    "sns",
    "ig",
    "instagram",
    "\u65c5\u8272ig",
    "\u65c5\u8272instagram",
    "\u304a\u53d6\u308a\u5bc4\u305b",
    "\u53f0\u6e7e",
)
SHEET_EXCLUDE_HINTS = (
    "fb",
    "tw",
    "line",
    "\u65bd\u7b56\u6574\u7406",
    "\u30c6\u30fc\u30de\u52df\u96c6",
    "\u30eb\u30fc\u30eb",
    "\u53c2\u7167",
    "\u30d5\u30a3\u30fc\u30c9\u30d0\u30c3\u30af",
)

HEADER_HINTS = (
    "\u6848\u4ef6\u540d",
    "\u5e97\u540d",
    "\u65bd\u8a2d\u540d",
    "\u30af\u30e9\u30a4\u30a2\u30f3\u30c8",
    "\u6295\u7a3f\u65e5",
    "\u516c\u958b\u6708",
    "\u30c6\u30fc\u30de",
    "\u65c5\u8272lp",
    "lp",
    "url",
    "\u62c5\u5f53",
    "\u62e0\u70b9",
    "\u30b8\u30e3\u30f3\u30eb",
)


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFKC", text)
    text = text.lower()
    return re.sub(r"\s+", "", text)


def stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def default_workbooks() -> list[Path]:
    downloads = Path.home() / "Downloads"
    return [downloads / name for name in DEFAULT_WORKBOOK_NAMES]


def sheet_is_relevant(title: str, include_all: bool = False) -> bool:
    if include_all:
        return True
    title_norm = norm(title)
    if any(hint in title_norm for hint in SHEET_EXCLUDE_HINTS):
        return False
    return any(hint in title_norm for hint in SHEET_INCLUDE_HINTS)


def account_guess(sheet_title: str) -> str:
    title_norm = norm(sheet_title)
    if "\u304a\u53d6\u308a\u5bc4\u305b" in title_norm:
        return "@tabiiro.otoriyose"
    if "\u53f0\u6e7e" in title_norm:
        return "@tabiiro_tw"
    if "\u8fd1\u757f" in title_norm or "kinki" in title_norm:
        return "@tabiiro.kinki"
    if "ig" in title_norm or "instagram" in title_norm or "sns" in title_norm:
        return "@tabiiro"
    return "unknown"


def best_header(recent_rows: Iterable[tuple[int, list[Any]]]) -> tuple[int | None, list[Any]]:
    best_score = 0
    best_row_num: int | None = None
    best_values: list[Any] = []
    for row_num, values in recent_rows:
        text = norm(" ".join(stringify(v) for v in values))
        non_empty = sum(1 for v in values if stringify(v))
        hints = sum(1 for hint in HEADER_HINTS if norm(hint) in text)
        if hints == 0 and non_empty < 4:
            continue
        score = hints * 10 + min(non_empty, 10)
        if score >= best_score:
            best_score = score
            best_row_num = row_num
            best_values = values
    return best_row_num, best_values


def row_context(values: list[Any], header_values: list[Any], keep_columns: set[int]) -> dict[str, str]:
    context: dict[str, str] = {}
    for idx, value in enumerate(values, start=1):
        if idx not in keep_columns:
            continue
        value_text = stringify(value)
        if not value_text:
            continue
        header = stringify(header_values[idx - 1]) if idx <= len(header_values) else ""
        key = header or get_column_letter(idx)
        if key in context:
            key = f"{key}_{get_column_letter(idx)}"
        context[key] = value_text
    return context


def search_workbook(path: Path, terms: list[str], include_all_sheets: bool) -> dict[str, Any]:
    result: dict[str, Any] = {
        "workbook": str(path),
        "exists": path.exists(),
        "searched_sheets": [],
        "matches": [],
    }
    if not path.exists():
        return result

    workbook = load_workbook(path, read_only=True, data_only=True)
    term_norms = [(term, norm(term)) for term in terms if norm(term)]

    for sheet in workbook.worksheets:
        if not sheet_is_relevant(sheet.title, include_all_sheets):
            continue
        result["searched_sheets"].append(sheet.title)
        recent: deque[tuple[int, list[Any]]] = deque(maxlen=60)
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = list(row)
            row_text = norm(" ".join(stringify(value) for value in values))
            matched = [term for term, term_norm in term_norms if term_norm and term_norm in row_text]
            if matched:
                header_row_num, header_values = best_header(recent)
                matched_cells = []
                matched_columns: set[int] = set()
                for col_index, value in enumerate(values, start=1):
                    value_text = stringify(value)
                    value_norm = norm(value_text)
                    cell_terms = [
                        term for term, term_norm in term_norms if term_norm and term_norm in value_norm
                    ]
                    if cell_terms:
                        matched_columns.add(col_index)
                        matched_cells.append(
                            {
                                "column": get_column_letter(col_index),
                                "value": value_text,
                                "matched_terms": cell_terms,
                            }
                        )
                keep_columns = set(range(1, min(len(values), 12) + 1))
                for col_index in matched_columns:
                    keep_columns.update(range(max(1, col_index - 3), min(len(values), col_index + 3) + 1))
                result["matches"].append(
                    {
                        "sheet": sheet.title,
                        "account_guess": account_guess(sheet.title),
                        "row": row_index,
                        "header_row": header_row_num,
                        "matched_terms": matched,
                        "matched_cells": matched_cells,
                        "row_context": row_context(values, header_values, keep_columns),
                    }
                )
            recent.append((row_index, values))

    return result


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--terms", nargs="+", required=True, help="Store-name variants, LP IDs, or slugs.")
    parser.add_argument("--workbooks", nargs="*", help="SNS workbook paths. Defaults to Downloads files.")
    parser.add_argument("--all-sheets", action="store_true", help="Search every sheet, including admin sheets.")
    parser.add_argument("--json", action="store_true", help="Emit JSON instead of a text summary.")
    args = parser.parse_args()

    paths = [Path(p) for p in args.workbooks] if args.workbooks else default_workbooks()
    workbook_results = [search_workbook(path, args.terms, args.all_sheets) for path in paths]
    matches = [match for workbook in workbook_results for match in workbook["matches"]]
    account_counts = Counter(match["account_guess"] for match in matches)

    output = {
        "workbooks": workbook_results,
        "terms": args.terms,
        "match_count": len(matches),
        "account_counts": dict(account_counts),
        "matches": matches,
    }

    if args.json:
        print(json.dumps(output, ensure_ascii=False, indent=2))
        return

    print(f"match_count: {len(matches)}")
    for match in matches:
        print(
            f"- {match['account_guess']} | {match['sheet']} row {match['row']} | "
            f"terms={', '.join(match['matched_terms'])}"
        )
        for key, value in list(match["row_context"].items())[:20]:
            print(f"  {key}: {value}")


if __name__ == "__main__":
    main()
