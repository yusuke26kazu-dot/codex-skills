#!/usr/bin/env python
"""Extract latest 3-month article metrics and query/rank rows from a workbook."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

MONTH_RE = re.compile(r"^\d{4}年\d{1,2}月号$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Path to .xlsx workbook")
    parser.add_argument("--terms", nargs="+", required=True, help="Article URL slugs or search terms")
    parser.add_argument(
        "--all-prefixes",
        action="store_true",
        help="Search every sheet even when sheets have numeric version prefixes",
    )
    parser.add_argument("--json", action="store_true", help="Emit JSON only")
    return parser.parse_args()


def text(value: Any) -> str:
    return "" if value is None else str(value)


def to_number(value: Any) -> float:
    if value in (None, "", "-"):
        return 0
    try:
        return float(value)
    except Exception:
        return 0


def sheet_prefix(title: str) -> int | None:
    match = re.match(r"^(\d+)【", title)
    return int(match.group(1)) if match else None


def find_header(rows: list[tuple[Any, ...]]) -> tuple[int, list[str], list[tuple[int, str]]] | None:
    for idx, row in enumerate(rows):
        header = [text(v) for v in row]
        month_cols = [(i, h) for i, h in enumerate(header) if MONTH_RE.match(h)]
        if month_cols:
            return idx, header, month_cols
    return None


def main() -> int:
    args = parse_args()
    workbook = Path(args.workbook)
    if not workbook.exists():
        print(f"Workbook not found: {workbook}", file=sys.stderr)
        return 2

    terms = [t for t in args.terms if t]
    wb = load_workbook(workbook, read_only=True, data_only=True)

    prefixes = [p for p in (sheet_prefix(ws.title) for ws in wb.worksheets) if p is not None]
    latest_prefix = max(prefixes) if prefixes and not args.all_prefixes else None

    results = []
    scanned_sheets = []

    for ws in wb.worksheets:
        prefix = sheet_prefix(ws.title)
        if latest_prefix is not None and prefix is not None and prefix != latest_prefix:
            continue
        if hasattr(ws, "reset_dimensions"):
            ws.reset_dimensions()
        rows = list(ws.iter_rows(values_only=True))
        header_info = find_header(rows)
        if not header_info:
            continue
        header_idx, header, month_cols = header_info
        if len(month_cols) < 3:
            continue

        last3 = month_cols[-3:]
        query_idx = header.index("取得クエリ") if "取得クエリ" in header else last3[-1][0] + 1
        rank_idx = header.index("順位") if "順位" in header else last3[-1][0] + 2
        scanned_sheets.append(ws.title)

        for row_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
            row_text = " ".join(text(v) for v in row)
            matched_terms = [term for term in terms if term in row_text]
            if not matched_terms:
                continue

            url_idx = next((i for i, v in enumerate(row) if any(term in text(v) for term in terms)), None)
            title = text(row[url_idx - 1]) if url_idx and url_idx - 1 >= 0 else ""
            url = text(row[url_idx]) if url_idx is not None else ""
            month_values = []
            for col_idx, label in last3:
                value = row[col_idx] if col_idx < len(row) else None
                number = to_number(value)
                month_values.append({"month": label, "value": int(number) if number.is_integer() else number})
            total = sum(float(item["value"]) for item in month_values)
            query = text(row[query_idx]) if query_idx < len(row) else ""
            rank_value = row[rank_idx] if rank_idx < len(row) else ""

            results.append(
                {
                    "sheet": ws.title,
                    "row": row_idx,
                    "title": title,
                    "url_or_term_cell": url,
                    "matched_terms": matched_terms,
                    "months": month_values,
                    "three_month_total": int(total) if float(total).is_integer() else total,
                    "query": query,
                    "rank": rank_value,
                }
            )

    grand_total = sum(float(r["three_month_total"]) for r in results)
    top10 = []
    for r in results:
        try:
            rank_num = float(r["rank"])
        except Exception:
            continue
        if rank_num <= 10:
            top10.append(r)

    payload = {
        "workbook": str(workbook),
        "terms": terms,
        "latest_prefix": latest_prefix,
        "scanned_sheets": scanned_sheets,
        "match_count": len(results),
        "grand_total": int(grand_total) if grand_total.is_integer() else grand_total,
        "results": results,
        "top10_keywords": [
            {
                "query": r["query"],
                "rank": r["rank"],
                "title": r["title"],
                "url_or_term_cell": r["url_or_term_cell"],
                "three_month_total": r["three_month_total"],
            }
            for r in top10
        ],
    }

    if args.json:
        print(json.dumps(payload, ensure_ascii=False, indent=2, default=str))
        return 0

    print(f"Workbook: {workbook}")
    if latest_prefix is not None:
        print(f"Latest sheet prefix: {latest_prefix}")
    print(f"Scanned sheets: {', '.join(scanned_sheets)}")
    print(f"Matches: {len(results)}")
    print(f"Grand total: {payload['grand_total']}")
    for r in results:
        months = ", ".join(f"{m['month']}={m['value']}" for m in r["months"])
        print(
            f"- {r['sheet']} row {r['row']}: {r['title']} | {months} | "
            f"total={r['three_month_total']} | query={r['query']} | rank={r['rank']}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
