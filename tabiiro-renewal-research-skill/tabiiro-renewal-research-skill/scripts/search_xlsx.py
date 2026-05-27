#!/usr/bin/env python
"""Search every cell in an Excel workbook for one or more terms."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Path to .xlsx workbook")
    parser.add_argument("--terms", nargs="+", required=True, help="Search terms")
    parser.add_argument("--context", type=int, default=0, help="Include N cells before/after in the same row")
    parser.add_argument("--json", action="store_true", help="Emit JSON only")
    return parser.parse_args()


def norm(value: object) -> str:
    return "" if value is None else str(value)


def main() -> int:
    args = parse_args()
    workbook = Path(args.workbook)
    if not workbook.exists():
        print(f"Workbook not found: {workbook}", file=sys.stderr)
        return 2

    terms = [t for t in args.terms if t]
    wb = load_workbook(workbook, read_only=True, data_only=True)
    matches = []

    for ws in wb.worksheets:
        if hasattr(ws, "reset_dimensions"):
            ws.reset_dimensions()
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_values = [norm(v) for v in row]
            for col_idx, text in enumerate(row_values, start=1):
                hit_terms = [term for term in terms if term in text]
                if not hit_terms:
                    continue
                start = max(1, col_idx - args.context)
                end = min(len(row_values), col_idx + args.context)
                matches.append(
                    {
                        "sheet": ws.title,
                        "row": row_idx,
                        "column": get_column_letter(col_idx),
                        "value": text,
                        "matched_terms": hit_terms,
                        "row_context": row_values[start - 1 : end] if args.context else [],
                    }
                )

    result = {
        "workbook": str(workbook),
        "sheet_count": len(wb.worksheets),
        "terms": terms,
        "match_count": len(matches),
        "matches": matches,
    }

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0

    print(f"Workbook: {workbook}")
    print(f"Sheets: {len(wb.worksheets)}")
    print(f"Terms: {', '.join(terms)}")
    print(f"Matches: {len(matches)}")
    for item in matches:
        print(
            f"- {item['sheet']}!{item['column']}{item['row']}: "
            f"{item['value']} (terms: {', '.join(item['matched_terms'])})"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
