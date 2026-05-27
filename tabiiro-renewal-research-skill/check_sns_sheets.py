import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
from openpyxl import load_workbook

downloads = Path.home() / "Downloads"
path = downloads / "【旅色SNS】営業支援投稿.xlsx"

if path.exists():
    wb = load_workbook(path, read_only=True)
    print("Sheets in SNS Excel:")
    for sheet in wb.sheetnames:
        print(" -", sheet)
else:
    print(f"File not found: {path}")
