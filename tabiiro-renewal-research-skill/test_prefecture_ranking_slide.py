import glob
import os
import re
import shutil
import win32com.client
import openpyxl

def extract_dsoco_rankings(excel_path):
    import datetime
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet3 = wb[wb.sheetnames[3]] # 都道府県別人気ランキング(グルメ)
    
    rankings = []
    rows = list(sheet3.iter_rows(values_only=True))
    total_rows = len(rows)
    
    for idx in range(total_rows):
        row = rows[idx]
        if not row:
            continue
            
        col0 = row[0]
        # Check if it's a date block header
        if isinstance(col0, datetime.datetime):
            date_val = col0
            # Format date as YYYY/M/D
            date_str = f"{date_val.year}/{date_val.month}/{date_val.day}"
            
            # Dynamically find the column for Osaka ('大阪' or '大阪府')
            osaka_col = None
            for c_idx in range(1, len(row)):
                c_val = str(row[c_idx]).strip() if row[c_idx] else ""
                if c_val == "大阪" or c_val == "大阪府":
                    osaka_col = c_idx
                    break
                    
            if osaka_col is None:
                continue
                
            # Inspect the next 5 rows (1位 to 5位)
            for offset in range(1, 6):
                next_idx = idx + offset
                if next_idx >= total_rows:
                    break
                next_row = rows[next_idx]
                if not next_row or len(next_row) <= osaka_col:
                    continue
                    
                rank_label = str(next_row[0]).strip() if next_row[0] else ""
                shop_name = str(next_row[osaka_col]).strip() if next_row[osaka_col] else ""
                
                if shop_name == "D-soco May's Cafe":
                    rankings.append({
                        "period": date_str,
                        "rank": rank_label
                    })
                    
    wb.close()
    return rankings

def replace_text_in_shapes(shapes, old_text, new_text):
    for shape in shapes:
        try:
            if shape.Type == 6: # Group shape
                replace_text_in_shapes(shape.GroupItems, old_text, new_text)
            elif shape.HasTextFrame and shape.TextFrame.HasText:
                txt = shape.TextFrame.TextRange.Text
                if old_text in txt:
                    shape.TextFrame.TextRange.Text = txt.replace(old_text, new_text)
        except Exception:
            pass

def generate_dsoco_ranking_slide():
    # 1. Resolve Excel file path dynamically
    excel_matches = glob.glob(r"C:\Users\NX023066\Downloads\*人気ランキング*.xlsx")
    excel_matches = [m for m in excel_matches if not os.path.basename(m).startswith("~$")]
    if not excel_matches:
        raise Exception("No popularity ranking Excel files found in Downloads!")
        
    excel_path = excel_matches[0]
    print(f"Resolved Excel Path: {excel_path}")
    
    # 2. Extract D-soco rankings
    rankings = extract_dsoco_rankings(excel_path)
    print(f"Extracted {len(rankings)} total rankings.")
    
    # 3. Duplicate template PPTX to target folder
    template_path = os.path.abspath(r"C:\Users\NX023066\Desktop\更新\案件ごと\施設報告資料\施設専用資料テンプレ（TG）260316.pptx")
    temp_pptx = os.path.abspath("temp_dsoco_ranking.pptx")
    shutil.copy2(template_path, temp_pptx)
    
    # 4. Open presentation in PowerPoint
    ppt = win32com.client.Dispatch("PowerPoint.Application")
    pres = ppt.Presentations.Open(temp_pptx, WithWindow=False)
    
    target_slide = None
    
    # 5. Delete all slides EXCEPT Slide 5 (Prefecture Ranking) backwards
    for i in range(pres.Slides.Count, 0, -1):
        slide = pres.Slides(i)
        text_content = ""
        def extract_text(shapes):
            nonlocal text_content
            for s in shapes:
                if s.Type == 6:
                    extract_text(s.GroupItems)
                elif s.HasTextFrame and s.TextFrame.HasText:
                    text_content += s.TextFrame.TextRange.Text + "\n"
        extract_text(slide.Shapes)
        
        # Identify the Prefecture Ranking slide
        if "ランクイン報告（都道府県別）" in text_content:
            target_slide = slide
            print(f"Keeping target Slide {i} ('ランクイン報告（都道府県別）')")
        else:
            slide.Delete()
            
    if not target_slide:
        pres.Close()
        ppt.Quit()
        if os.path.exists(temp_pptx):
            os.remove(temp_pptx)
        raise Exception("Prefecture Ranking slide was not found in the template!")
        
    # 6. Apply text replacements in shapes
    # Upper text box: ●●県 -> 大阪府
    replace_text_in_shapes(target_slide.Shapes, "●●県", "大阪府")
    # Lower text box: ○○県 -> 大阪府
    replace_text_in_shapes(target_slide.Shapes, "○○県", "大阪府")
    # Delete "●回" from the lower text box
    replace_text_in_shapes(target_slide.Shapes, "●回", "")
    
    # 7. Identify and sort Left and Right tables dynamically by Left position
    table_shapes = [s for s in target_slide.Shapes if s.HasTable]
    if len(table_shapes) != 2:
        pres.Close()
        ppt.Quit()
        if os.path.exists(temp_pptx):
            os.remove(temp_pptx)
        raise Exception(f"Expected exactly 2 tables on the slide, found {len(table_shapes)}")
        
    table_shapes.sort(key=lambda s: s.Left)
    left_table = table_shapes[0].Table
    right_table = table_shapes[1].Table
    
    # 8. Align and fill rankings backwards from the bottom of the right table
    total_slots = 14
    kept_rankings = rankings[-total_slots:]
    
    print(f"Filling {len(kept_rankings)} most recent rankings into {total_slots} available slots...")
    
    for i in range(total_slots):
        idx = i - (total_slots - len(kept_rankings))
        if idx >= 0:
            period = kept_rankings[idx]["period"]
            rank = kept_rankings[idx]["rank"]
        else:
            period = ""
            rank = ""
            
        # Target table and row (rows 2 to 8 are the 7 data rows)
        if i < 7:
            row_num = i + 2
            left_table.Cell(row_num, 1).Shape.TextFrame.TextRange.Text = period
            left_table.Cell(row_num, 2).Shape.TextFrame.TextRange.Text = rank
            
            # Ensure text is center-aligned and has premium font sizes
            left_table.Cell(row_num, 1).Shape.TextFrame.TextRange.ParagraphFormat.Alignment = 2
            left_table.Cell(row_num, 2).Shape.TextFrame.TextRange.ParagraphFormat.Alignment = 2
            left_table.Cell(row_num, 1).Shape.TextFrame.TextRange.Font.Size = 11
            left_table.Cell(row_num, 2).Shape.TextFrame.TextRange.Font.Size = 11
        else:
            row_num = (i - 7) + 2
            right_table.Cell(row_num, 1).Shape.TextFrame.TextRange.Text = period
            right_table.Cell(row_num, 2).Shape.TextFrame.TextRange.Text = rank
            
            # Ensure text is center-aligned and has premium font sizes
            right_table.Cell(row_num, 1).Shape.TextFrame.TextRange.ParagraphFormat.Alignment = 2
            right_table.Cell(row_num, 2).Shape.TextFrame.TextRange.ParagraphFormat.Alignment = 2
            right_table.Cell(row_num, 1).Shape.TextFrame.TextRange.Font.Size = 11
            right_table.Cell(row_num, 2).Shape.TextFrame.TextRange.Font.Size = 11
            
    # 9. Save presentation and move to Google Drive
    pres.Save()
    pres.Close()
    ppt.Quit()
    
    output_path = r"G:\マイドライブ\codex-skills\tbiiro-renewal\更新資料\test_prefecture_ranking.pptx"
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except Exception:
            pass
            
    shutil.move(temp_pptx, output_path)
    print(f"Prefecture ranking slide successfully created at: {output_path}")

if __name__ == "__main__":
    generate_dsoco_ranking_slide()
